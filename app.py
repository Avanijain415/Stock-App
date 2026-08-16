import os
import sqlite3
from datetime import datetime, timezone, timedelta
from flask import Flask, render_template, request, jsonify, session, Response
import openpyxl

app = Flask(__name__)
app.secret_key = 'enterprise_super_secret_key_2026'

DB_NAME = 'stock.db'
EXCEL_FILE = 'NEW APRIL STOCK SHEET 2026.xlsx'
MASTER_RESET_PIN = '2026'

IST = timezone(timedelta(hours=5, minutes=30))

def get_current_ist_time():
    return datetime.now(IST).strftime('%Y-%m-%d %I:%M:%S %p')

def get_current_ist_date():
    return datetime.now(IST).strftime('%Y-%m-%d')

def get_db():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT NOT NULL
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            hsn_code TEXT,
            country TEXT,
            brand TEXT,
            description TEXT NOT NULL,
            price REAL DEFAULT 0,
            tax REAL DEFAULT 0,
            price_after_tax REAL DEFAULT 0,
            mrp REAL DEFAULT 0,
            per_petti INTEGER DEFAULT 0,
            expiry_date TEXT,
            opening_stock INTEGER DEFAULT 0,
            actual_stock INTEGER DEFAULT 0,
            damage INTEGER DEFAULT 0,
            status TEXT DEFAULT 'Completed'
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS marketplace_products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            marketplace TEXT NOT NULL,
            marketplace_label TEXT NOT NULL,
            currency_symbol TEXT DEFAULT '$',
            sku TEXT,
            asin TEXT,
            selling_price REAL DEFAULT 0,
            fba_fee REAL DEFAULT 0,
            purchase_price REAL DEFAULT 0,
            conv_rate REAL DEFAULT 1,
            purchase_rate REAL DEFAULT 0,
            freight REAL DEFAULT 0,
            min_selling_rate REAL DEFAULT 0,
            margin REAL DEFAULT 0,
            product_link TEXT
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS activity_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER,
            product_desc TEXT,
            username TEXT,
            action TEXT,
            quantity INTEGER,
            prev_stock INTEGER,
            new_stock INTEGER,
            details TEXT,
            timestamp TEXT
        )
    ''')
    
    cursor.execute("INSERT OR IGNORE INTO users (username, password, role) VALUES ('master', 'admin123', 'master')")
    cursor.execute("INSERT OR IGNORE INTO users (username, password, role) VALUES ('worker1', 'worker123', 'worker')")
    cursor.execute("INSERT OR IGNORE INTO users (username, password, role) VALUES ('worker2', 'worker123', 'worker')")
    
    conn.commit()
    conn.close()

def migrate_from_excel():
    conn = get_db()
    cursor = conn.cursor()
    
    if not os.path.exists(EXCEL_FILE):
        conn.close()
        return

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    
    # 1. Migrate Warehouse Products with True Running Balance Math
    cursor.execute("SELECT COUNT(*) FROM products")
    if cursor.fetchone()[0] == 0 and 'Sheet1' in wb.sheetnames:
        sheet = wb['Sheet1']
        for row in range(3, sheet.max_row + 1):
            desc = sheet.cell(row=row, column=5).value
            if not desc or str(desc).strip().lower() in ['discription', 'description', 'none', '']:
                continue
                
            hsn = str(sheet.cell(row=row, column=2).value or '').replace('.0', '')
            country = str(sheet.cell(row=row, column=3).value or 'INDIA')
            brand = str(sheet.cell(row=row, column=4).value or 'GENERAL')
            
            try: price = float(sheet.cell(row=row, column=6).value or 0)
            except: price = 0.0
            try: tax = float(sheet.cell(row=row, column=7).value or 0)
            except: tax = 0.0
            try: 
                price_after_tax = float(sheet.cell(row=row, column=8).value or (price * (1 + tax)))
            except: 
                price_after_tax = price
            try: mrp = float(sheet.cell(row=row, column=9).value or 0)
            except: mrp = 0.0
            try: per_petti = int(sheet.cell(row=row, column=10).value or 0)
            except: per_petti = 0
            
            expiry = str(sheet.cell(row=row, column=11).value or '')
            try: opening = int(float(sheet.cell(row=row, column=14).value or 0))
            except: opening = 0
            try: actual = int(float(sheet.cell(row=row, column=15).value or opening))
            except: actual = opening
            try: damage = int(float(sheet.cell(row=row, column=17).value or 0))
            except: damage = 0

            status = "Pending" if actual <= 5 else "Completed"

            cursor.execute('''
                INSERT INTO products (hsn_code, country, brand, description, price, tax, price_after_tax, mrp, per_petti, expiry_date, opening_stock, actual_stock, damage, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (hsn, country, brand, str(desc).strip(), price, tax, price_after_tax, mrp, per_petti, expiry, opening, actual, damage, status))
            prod_id = cursor.lastrowid
            
            # Step-by-Step Running Balance Ledger for Historical Entries
            running_balance = opening
            for col in range(18, min(sheet.max_column + 1, 50)):
                val = sheet.cell(row=row, column=col).value
                col_hdr = sheet.cell(row=2, column=col).value
                if val is not None and str(val).strip() not in ['', 'None', 'nan']:
                    try:
                        qty = int(float(val))
                        if qty == 0: continue
                        
                        action = "RECEIVE" if qty > 0 else "ISSUE"
                        prev_bal = running_balance
                        running_balance = max(0, running_balance + qty)
                        date_str = str(col_hdr)[:10] if col_hdr else "Excel Log"
                        
                        cursor.execute('''
                            INSERT INTO activity_log (product_id, product_desc, username, action, quantity, prev_stock, new_stock, details, timestamp)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ''', (prod_id, str(desc).strip(), 'System Excel', action, abs(qty), prev_bal, running_balance, f"Initial Data from {date_str}", get_current_ist_time()))
                    except:
                        pass

    # 2. Migrate Global Country Marketplace Sheets with Guaranteed Amazon Fallback URLs
    cursor.execute("SELECT COUNT(*) FROM marketplace_products")
    if cursor.fetchone()[0] == 0:
        marketplaces = [
            ('GTUS', 'USA 🇺🇸 (GTUS)', '$', 95.0, 'https://www.amazon.com/dp/'),
            ('CANADA', 'Canada 🇨🇦', 'CA$', 66.0, 'https://www.amazon.ca/dp/'),
            ('GTAU', 'Australia 🇦🇺 (GTAU)', 'A$', 65.0, 'https://www.amazon.com.au/dp/'),
            ('UAE', 'UAE 🇦🇪', 'AED ', 25.0, 'https://www.amazon.ae/dp/'),
            ('INDIA', 'India 🇮🇳', '₹', 1.0, 'https://www.amazon.in/dp/'),
            ('VAIS NEW', 'USA 🇺🇸 (VAIS)', '$', 94.0, 'https://www.amazon.com/dp/'),
            ('RTAU', 'Australia 🇦🇺 (RTAU)', 'A$', 65.0, 'https://www.amazon.com.au/dp/')
        ]
        
        for m_sheet, m_label, m_curr, default_conv, amazon_base in marketplaces:
            if m_sheet not in wb.sheetnames:
                continue
            ws = wb[m_sheet]
            
            header_row = None
            header_map = {}
            for r in range(1, 10):
                row_vals = [str(ws.cell(row=r, column=c).value or '').strip().upper() for c in range(1, 16)]
                if any('SKU' in v for v in row_vals):
                    header_row = r
                    for c_idx, val in enumerate(row_vals):
                        if val:
                            header_map[val] = c_idx + 1
                    break
            
            if not header_row:
                continue
            
            def get_val_by_keys(row_idx, keys, default=0.0):
                for k in keys:
                    for h_name, col_num in header_map.items():
                        if k in h_name:
                            v = ws.cell(row=row_idx, column=col_num).value
                            if v is not None:
                                try: return float(v)
                                except: return default
                return default

            for r in range(header_row + 1, ws.max_row + 1):
                sku_col = header_map.get('SKU', header_map.get('SELLER-SKU', 2))
                sku = ws.cell(row=r, column=sku_col).value
                if not sku or str(sku).strip().lower() in ['none', '', 'sku', 'seller-sku']:
                    continue
                
                asin_col = header_map.get('ASIN', 3)
                asin = str(ws.cell(row=r, column=asin_col).value or '').strip().replace('\n', '')
                
                price = get_val_by_keys(r, ['PRICE'])
                fba = get_val_by_keys(r, ['FBA FEE', 'FBA'])
                purchase_price = get_val_by_keys(r, ['PURCHASE PRICE', 'PURCHASE RATE'])
                conv_rate = get_val_by_keys(r, ['CONV RATE', 'CONVERSION RATE'], default=default_conv)
                if conv_rate <= 0: conv_rate = default_conv
                
                freight = get_val_by_keys(r, ['FREIGHT', 'FRIEGHT'])
                min_selling = get_val_by_keys(r, ['MIN SELLING RATE', 'MIN SELLING'])
                margin = get_val_by_keys(r, ['MARGIN'])
                
                # Check for explicit link in sheet or create guaranteed Amazon URL via ASIN
                link = ""
                for c in range(1, 16):
                    cell_v = str(ws.cell(row=r, column=c).value or '').strip()
                    if cell_v.startswith('http'):
                        link = cell_v
                        break
                
                if not link and asin and len(asin) >= 5:
                    link = f"{amazon_base}{asin}"
                
                cursor.execute('''
                    INSERT INTO marketplace_products 
                    (marketplace, marketplace_label, currency_symbol, sku, asin, selling_price, fba_fee, purchase_price, conv_rate, purchase_rate, freight, min_selling_rate, margin, product_link)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (m_sheet, m_label, m_curr, str(sku).strip(), asin, price, fba, purchase_price, conv_rate, purchase_price, freight, min_selling, margin, link))

    conn.commit()
    conn.close()

if os.path.exists(DB_NAME):
    try: os.remove(DB_NAME)
    except: pass

init_db()
migrate_from_excel()

# ----------------- APIS -----------------

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/api/login', methods=['POST'])
def login():
    data = request.json or {}
    u = data.get('username', '').strip()
    p = data.get('password', '').strip()
    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE username = ? AND password = ?", (u, p)).fetchone()
    conn.close()
    if user:
        session['username'] = user['username']
        session['role'] = user['role']
        return jsonify({"status": "success", "username": user['username'], "role": user['role']})
    return jsonify({"status": "error", "message": "Invalid username or password"}), 401

@app.route('/api/reset-password', methods=['POST'])
def reset_password():
    data = request.json or {}
    username = data.get('username', '').strip()
    new_password = data.get('new_password', '').strip()
    pin = data.get('pin', '').strip()
    
    if pin != MASTER_RESET_PIN:
        return jsonify({"status": "error", "message": "Invalid Master Recovery PIN"}), 403
    
    if not username or not new_password:
        return jsonify({"status": "error", "message": "All fields are required"}), 400

    conn = get_db()
    cursor = conn.cursor()
    user = cursor.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
    if not user:
        conn.close()
        return jsonify({"status": "error", "message": "User not found"}), 404
        
    cursor.execute("UPDATE users SET password = ? WHERE username = ?", (new_password, username))
    conn.commit()
    conn.close()
    return jsonify({"status": "success", "message": "Password reset successfully!"})

@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({"status": "success"})

@app.route('/api/worker-stats', methods=['GET'])
def worker_stats():
    username = session.get('username', '')
    today_ist = get_current_ist_date()
    conn = get_db()
    today_received = conn.execute("SELECT SUM(quantity) FROM activity_log WHERE username = ? AND action = 'RECEIVE' AND timestamp LIKE ?", (username, f"{today_ist}%")).fetchone()[0] or 0
    today_issued = conn.execute("SELECT SUM(quantity) FROM activity_log WHERE username = ? AND action = 'ISSUE' AND timestamp LIKE ?", (username, f"{today_ist}%")).fetchone()[0] or 0
    today_ops = conn.execute("SELECT COUNT(*) FROM activity_log WHERE username = ? AND timestamp LIKE ?", (username, f"{today_ist}%")).fetchone()[0] or 0
    conn.close()
    return jsonify({
        "today_received": today_received,
        "today_issued": today_issued,
        "today_ops": today_ops
    })

@app.route('/api/stats', methods=['GET'])
def get_stats():
    conn = get_db()
    total_items = conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]
    total_qty = conn.execute("SELECT SUM(actual_stock) FROM products").fetchone()[0] or 0
    low_stock = conn.execute("SELECT COUNT(*) FROM products WHERE actual_stock <= 10").fetchone()[0]
    pending_tasks = conn.execute("SELECT COUNT(*) FROM products WHERE status = 'Pending'").fetchone()[0]
    total_global_skus = conn.execute("SELECT COUNT(*) FROM marketplace_products").fetchone()[0]
    
    all_prods = conn.execute("SELECT actual_stock, damage, price_after_tax, price FROM products").fetchall()
    total_inventory_val = 0.0
    total_damage_loss = 0.0
    for p in all_prods:
        unit_p = float(p['price_after_tax'] or p['price'] or 0)
        total_inventory_val += float(p['actual_stock'] or 0) * unit_p
        total_damage_loss += float(p['damage'] or 0) * unit_p
    
    top_stocks = conn.execute("SELECT description, actual_stock FROM products ORDER BY actual_stock DESC LIMIT 6").fetchall()
    brand_breakdown = conn.execute("SELECT brand, COUNT(*), SUM(actual_stock) FROM products GROUP BY brand ORDER BY SUM(actual_stock) DESC LIMIT 5").fetchall()
    marketplace_breakdown = conn.execute("SELECT marketplace_label, COUNT(*), AVG(margin) FROM marketplace_products GROUP BY marketplace_label").fetchall()
    
    conn.close()
    return jsonify({
        "total_items": total_items,
        "total_qty": total_qty,
        "low_stock": low_stock,
        "pending_tasks": pending_tasks,
        "total_global_skus": total_global_skus,
        "total_inventory_val": round(total_inventory_val, 2),
        "total_damage_loss": round(total_damage_loss, 2),
        "top_stocks": [dict(r) for r in top_stocks],
        "brand_breakdown": [{"brand": r[0], "count": r[1], "qty": r[2] or 0} for r in brand_breakdown],
        "marketplace_breakdown": [{"label": r[0], "skus": r[1], "avg_margin": round(r[2] or 0, 2)} for r in marketplace_breakdown]
    })

@app.route('/api/products', methods=['GET'])
def get_products():
    conn = get_db()
    products = conn.execute("SELECT * FROM products ORDER BY id ASC").fetchall()
    conn.close()
    return jsonify([dict(p) for p in products])

@app.route('/api/products/add', methods=['POST'])
def add_product():
    if 'username' not in session:
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
    
    data = request.json or {}
    desc = data.get('description', '').strip()
    if not desc:
        return jsonify({"status": "error", "message": "Product description is required"}), 400
    
    hsn = data.get('hsn_code', '').strip()
    brand = data.get('brand', 'GENERAL').strip()
    country = data.get('country', 'INDIA').strip()
    price = float(data.get('price', 0) or 0)
    tax = float(data.get('tax', 0.18) or 0.18)
    price_after_tax = round(price * (1 + tax), 2)
    mrp = float(data.get('mrp', 0) or 0)
    per_petti = int(data.get('per_petti', 0) or 0)
    expiry = data.get('expiry_date', '').strip()
    opening = int(data.get('opening_stock', 0) or 0)
    status = "Pending" if opening <= 5 else "Completed"
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO products (hsn_code, country, brand, description, price, tax, price_after_tax, mrp, per_petti, expiry_date, opening_stock, actual_stock, damage, status)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?)
    ''', (hsn, country, brand, desc, price, tax, price_after_tax, mrp, per_petti, expiry, opening, opening, status))
    prod_id = cursor.lastrowid
    
    cursor.execute('''
        INSERT INTO activity_log (product_id, product_desc, username, action, quantity, prev_stock, new_stock, details, timestamp)
        VALUES (?, ?, ?, 'NEW_ITEM', ?, 0, ?, 'Product Added to Catalog', ?)
    ''', (prod_id, desc, session['username'], opening, opening, get_current_ist_time()))
    
    conn.commit()
    conn.close()
    return jsonify({"status": "success", "message": "Product added successfully!"})

@app.route('/api/products/delete/<int:prod_id>', methods=['DELETE'])
def delete_product(prod_id):
    if session.get('role') != 'master':
        return jsonify({"status": "error", "message": "Forbidden: Master authority required"}), 403
    
    conn = get_db()
    cursor = conn.cursor()
    prod = cursor.execute("SELECT * FROM products WHERE id = ?", (prod_id,)).fetchone()
    if not prod:
        conn.close()
        return jsonify({"status": "error", "message": "Product not found"}), 404
        
    cursor.execute("DELETE FROM products WHERE id = ?", (prod_id,))
    cursor.execute('''
        INSERT INTO activity_log (product_id, product_desc, username, action, quantity, prev_stock, new_stock, details, timestamp)
        VALUES (?, ?, ?, 'DELETE', 0, ?, 0, 'Product Removed by Master', ?)
    ''', (prod_id, prod['description'], session['username'], prod['actual_stock'], get_current_ist_time()))
    
    conn.commit()
    conn.close()
    return jsonify({"status": "success", "message": f"Product #{prod_id} successfully deleted"})

@app.route('/api/products/edit', methods=['POST'])
def edit_product():
    if session.get('role') != 'master':
        return jsonify({"status": "error", "message": "Forbidden: Master authority required"}), 403
        
    data = request.json or {}
    prod_id = data.get('id')
    desc = data.get('description', '').strip()
    if not desc:
        return jsonify({"status": "error", "message": "Description cannot be empty"}), 400
        
    hsn = data.get('hsn_code', '').strip()
    brand = data.get('brand', 'GENERAL').strip()
    country = data.get('country', 'INDIA').strip()
    price = float(data.get('price', 0) or 0)
    mrp = float(data.get('mrp', 0) or 0)
    tax = float(data.get('tax', 0.18) or 0.18)
    price_after_tax = round(price * (1 + tax), 2)
    per_petti = int(data.get('per_petti', 0) or 0)
    expiry = data.get('expiry_date', '').strip()
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE products SET 
            description = ?, hsn_code = ?, brand = ?, country = ?, 
            price = ?, tax = ?, price_after_tax = ?, mrp = ?, 
            per_petti = ?, expiry_date = ?
        WHERE id = ?
    ''', (desc, hsn, brand, country, price, tax, price_after_tax, mrp, per_petti, expiry, prod_id))
    
    conn.commit()
    conn.close()
    return jsonify({"status": "success", "message": "Product updated successfully!"})

@app.route('/api/products/adjust', methods=['POST'])
def adjust_stock():
    if 'username' not in session:
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
    
    data = request.json or {}
    prod_id = data.get('id')
    mode = data.get('mode', 'STOCK')
    change = int(data.get('change', 0))
    remarks = data.get('remarks', '').strip()
    
    conn = get_db()
    cursor = conn.cursor()
    prod = cursor.execute("SELECT * FROM products WHERE id = ?", (prod_id,)).fetchone()
    if not prod:
        conn.close()
        return jsonify({"status": "error", "message": "Product not found"}), 404
    
    prev_stock = prod['actual_stock']
    prev_damage = prod['damage']
    
    if mode == 'DAMAGE':
        if change > prev_stock:
            conn.close()
            return jsonify({"status": "error", "message": f"Cannot mark {change} units as damaged. Only {prev_stock} available!"}), 400
        new_stock = prev_stock - change
        new_damage = prev_damage + change
        status = "Pending" if new_stock <= 5 else "Completed"
        cursor.execute("UPDATE products SET actual_stock = ?, damage = ?, status = ? WHERE id = ?", (new_stock, new_damage, status, prod_id))
        detail_str = remarks if remarks else f"Moved {change} units to Damaged/Defective"
        cursor.execute('''
            INSERT INTO activity_log (product_id, product_desc, username, action, quantity, prev_stock, new_stock, details, timestamp)
            VALUES (?, ?, ?, 'DAMAGE', ?, ?, ?, ?, ?)
        ''', (prod_id, prod['description'], session['username'], change, prev_stock, new_stock, detail_str, get_current_ist_time()))
        log_id = cursor.lastrowid
    else:
        action_type = "RECEIVE" if change > 0 else "ISSUE"
        if change < 0 and abs(change) > prev_stock:
            conn.close()
            return jsonify({"status": "error", "message": f"Insufficient Stock! Available: {prev_stock}, Requested: {abs(change)}"}), 400
            
        new_stock = max(0, prev_stock + change)
        status = "Pending" if new_stock <= 5 else "Completed"
        cursor.execute("UPDATE products SET actual_stock = ?, status = ? WHERE id = ?", (new_stock, status, prod_id))
        detail_str = remarks if remarks else f"{session['username']} {action_type}D {abs(change)} units"
        cursor.execute('''
            INSERT INTO activity_log (product_id, product_desc, username, action, quantity, prev_stock, new_stock, details, timestamp)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (prod_id, prod['description'], session['username'], action_type, abs(change), prev_stock, new_stock, detail_str, get_current_ist_time()))
        log_id = cursor.lastrowid
    
    conn.commit()
    conn.close()
    return jsonify({"status": "success", "new_stock": new_stock, "log_id": log_id, "prev_stock": prev_stock, "prod_id": prod_id})

@app.route('/api/products/undo', methods=['POST'])
def undo_adjustment():
    if 'username' not in session:
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
    
    data = request.json or {}
    log_id = data.get('log_id')
    conn = get_db()
    cursor = conn.cursor()
    log = cursor.execute("SELECT * FROM activity_log WHERE id = ?", (log_id,)).fetchone()
    if not log:
        conn.close()
        return jsonify({"status": "error", "message": "Transaction log not found"}), 404
        
    prod_id = log['product_id']
    prev_stock = log['prev_stock']
    status = "Pending" if prev_stock <= 5 else "Completed"
    
    cursor.execute("UPDATE products SET actual_stock = ?, status = ? WHERE id = ?", (prev_stock, status, prod_id))
    cursor.execute('''
        INSERT INTO activity_log (product_id, product_desc, username, action, quantity, prev_stock, new_stock, details, timestamp)
        VALUES (?, ?, ?, 'UNDO', ?, ?, ?, 'Undo Action Applied', ?)
    ''', (prod_id, log['product_desc'], session['username'], log['quantity'], log['new_stock'], prev_stock, get_current_ist_time()))
    
    conn.commit()
    conn.close()
    return jsonify({"status": "success", "message": "Action successfully reverted!"})

@app.route('/api/marketplaces/products', methods=['GET'])
def get_marketplace_products():
    if session.get('role') != 'master':
        return jsonify({"status": "error", "message": "Confidential financial data reserved for Master role"}), 403
    mp = request.args.get('marketplace', 'GTUS')
    conn = get_db()
    items = conn.execute("SELECT * FROM marketplace_products WHERE marketplace = ?", (mp,)).fetchall()
    conn.close()
    return jsonify([dict(i) for i in items])

@app.route('/api/ledger', methods=['GET'])
def get_ledger():
    prod_id = request.args.get('product_id')
    action_type = request.args.get('action')
    conn = get_db()
    query = "SELECT * FROM activity_log"
    params = []
    conditions = []
    
    if prod_id:
        conditions.append("product_id = ?")
        params.append(prod_id)
    if action_type and action_type != 'ALL':
        conditions.append("action = ?")
        params.append(action_type)
        
    if conditions:
        query += " WHERE " + " AND ".join(conditions)
        
    query += " ORDER BY id DESC LIMIT 150"
    logs = conn.execute(query, params).fetchall()
    conn.close()
    return jsonify([dict(l) for l in logs])

# ----------------- CSV EXPORT SUITE -----------------

@app.route('/api/export/warehouse-csv')
def export_warehouse_csv():
    conn = get_db()
    products = conn.execute("SELECT * FROM products ORDER BY id ASC").fetchall()
    conn.close()
    
    def generate():
        yield "ID,HSN Code,Country,Brand,Description,Price,Tax,Price After Tax,MRP,Units Per Petti,Expiry Date,Opening Stock,Actual Stock,Damage Units,Status\n"
        for p in products:
            desc = p["description"].replace('"', '""')
            yield f'"{p["id"]}","{p["hsn_code"]}","{p["country"]}","{p["brand"]}","{desc}","{p["price"]}","{p["tax"]}","{p["price_after_tax"]}","{p["mrp"]}","{p["per_petti"]}","{p["expiry_date"]}","{p["opening_stock"]}","{p["actual_stock"]}","{p["damage"]}","{p["status"]}"\n'
            
    return Response(generate(), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=Warehouse_Stock_Report.csv"})

@app.route('/api/export/marketplace-csv')
def export_marketplace_csv():
    if session.get('role') != 'master':
        return jsonify({"status": "error", "message": "Forbidden"}), 403
    mp = request.args.get('marketplace', 'GTUS')
    conn = get_db()
    items = conn.execute("SELECT * FROM marketplace_products WHERE marketplace = ?", (mp,)).fetchall()
    conn.close()
    
    def generate():
        yield "ID,Marketplace,Currency,SKU,ASIN,Selling Price,FBA Fee,Purchase Price,Conversion Rate,Purchase Rate,Freight,Min Selling Price,Profit Margin,Product Link\n"
        for item in items:
            sku = item["sku"].replace('"', '""')
            yield f'"{item["id"]}","{item["marketplace_label"]}","{item["currency_symbol"]}","{sku}","{item["asin"]}","{item["selling_price"]}","{item["fba_fee"]}","{item["purchase_price"]}","{item["conv_rate"]}","{item["purchase_rate"]}","{item["freight"]}","{item["min_selling_rate"]}","{item["margin"]}","{item["product_link"]}"\n'
            
    return Response(generate(), mimetype="text/csv", headers={"Content-Disposition": f"attachment; filename=Marketplace_{mp}_Pricing_Report.csv"})

@app.route('/api/export/ledger-csv')
def export_ledger_csv():
    prod_id = request.args.get('product_id', '')
    conn = get_db()
    if prod_id:
        logs = conn.execute("SELECT * FROM activity_log WHERE product_id = ? ORDER BY id DESC", (prod_id,)).fetchall()
    else:
        logs = conn.execute("SELECT * FROM activity_log ORDER BY id DESC").fetchall()
    conn.close()
    
    def generate():
        yield "ID,Product ID,Product Name,Operator,Action,Quantity,Previous Stock,New Balance,Audit Details,Timestamp (IST)\n"
        for l in logs:
            pname = l["product_desc"].replace('"', '""')
            det = (l["details"] or "").replace('"', '""')
            yield f'"{l["id"]}","{l["product_id"]}","{pname}","{l["username"]}","{l["action"]}","{l["quantity"]}","{l["prev_stock"]}","{l["new_stock"]}","{det}","{l["timestamp"]}"\n'
            
    return Response(generate(), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=Complete_Audit_Ledger.csv"})

if __name__ == '__main__':
    app.run(debug=True, port=5000)
